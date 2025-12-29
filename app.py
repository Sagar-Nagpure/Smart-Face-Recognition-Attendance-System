import os
import cv2
import numpy as np
import pandas as pd
import time
from flask import Flask, request, render_template, redirect, url_for, session, flash
from datetime import datetime, date
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook, load_workbook
from mtcnn.mtcnn import MTCNN
from keras_facenet import FaceNet
import joblib
import threading
from concurrent.futures import ThreadPoolExecutor

# Initialize Flask and models
app = Flask(__name__)
app.secret_key = os.urandom(24)  # More secure secret key

# Optimized detector with better parameters
detector = MTCNN(steps_threshold=[0.7, 0.8, 0.9])
embedder = FaceNet()

# Paths
ATTENDANCE_DIR = 'Attendance'
FACES_DIR = 'static/faces'
MODEL_PATH = 'static/face_recognition_model.pkl'
os.makedirs(ATTENDANCE_DIR, exist_ok=True)
os.makedirs(FACES_DIR, exist_ok=True)

# Global variables
MESSAGE = "Press 'a' to mark attendance or 'q' to quit."
TODAY = date.today().strftime("%m_%d_%y")
CSV_PATH = os.path.join(ATTENDANCE_DIR, f'Attendance-{TODAY}.csv')
EXCEL_PATH = os.path.join(ATTENDANCE_DIR, f'Attendance-{TODAY}.xlsx')

# Thread pool for parallel processing
executor = ThreadPoolExecutor(max_workers=4)

# Initialize attendance files
def init_attendance_files():
    if not os.path.isfile(CSV_PATH):
        pd.DataFrame(columns=['Name', 'Roll', 'Time']).to_csv(CSV_PATH, index=False)
    
    if not os.path.isfile(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Roll', 'Time'])
        wb.save(EXCEL_PATH)

# Initialize files on app start
init_attendance_files()

# ----------- Helper Functions -----------

def get_embedding(face):
    face = cv2.resize(face, (160, 160))
    return embedder.embeddings([face])[0]

def extract_faces_mtcnn(img):
    # Convert to RGB for MTCNN
    rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    results = detector.detect_faces(rgb_img)
    faces, boxes = [], []
    h, w = img.shape[:2]
    for res in results:
        x, y, width, height = res['box']
        x, y = max(0, x), max(0, y)
        x2, y2 = min(w, x + width), min(h, y + height)
        face = img[y:y2, x:x2]
        if face.size == 0: continue
        faces.append(face)
        boxes.append((x, y, x2 - x, y2 - y))
    return faces, boxes

def is_blurry(face, threshold=30.0):
    if len(face.shape) == 3:
        gray = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
    else:
        gray = face
    return cv2.Laplacian(gray, cv2.CV_64F).var() < threshold

def train_model():
    """Train the face recognition model in a background thread"""
    def train_task():
        embeddings, labels = [], []
        for user in os.listdir(FACES_DIR):
            user_path = os.path.join(FACES_DIR, user)
            for imgname in os.listdir(user_path):
                img_path = os.path.join(user_path, imgname)
                img = cv2.imread(img_path)
                if img is None: continue
                try:
                    emb = get_embedding(img)
                    embeddings.append(emb)
                    labels.append(user)
                except Exception as e:
                    print(f"Error processing {img_path}: {str(e)}")
                    continue
        if embeddings:
            joblib.dump((embeddings, labels), MODEL_PATH)
            print("Model trained successfully!")
    
    thread = threading.Thread(target=train_task)
    thread.start()

def recognize_face(face):
    if not os.path.exists(MODEL_PATH):
        return "Unknown"
    embedding = get_embedding(face)
    embeddings, labels = joblib.load(MODEL_PATH)
    sims = cosine_similarity([embedding], embeddings)[0]
    max_idx = np.argmax(sims)
    max_score = sims[max_idx]
    if max_score >= 0.70:  # Increased threshold for better accuracy
        return labels[max_idx]
    return "Unknown"

def add_attendance(name):
    username, roll = name.split('_')
    now = datetime.now().strftime("%H:%M:%S")
    
    # Add to CSV
    df_csv = pd.read_csv(CSV_PATH)
    if not ((df_csv['Name'] == username) & (df_csv['Roll'] == roll)).any():
        new_entry = pd.DataFrame([[username, roll, now]], 
                                 columns=['Name', 'Roll', 'Time'])
        new_entry.to_csv(CSV_PATH, mode='a', header=False, index=False)
    
    # Add to Excel
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.append([username, roll, now])
    wb.save(EXCEL_PATH)

def extract_attendance():
    return pd.read_csv(CSV_PATH)

def get_registered_users():
    names, rolls, times = [], [], []
    for user in os.listdir(FACES_DIR):
        try:
            name, roll = user.split('_')
            names.append(name)
            rolls.append(roll)
            # Get registration time from folder creation time
            user_path = os.path.join(FACES_DIR, user)
            ctime = os.path.getctime(user_path)
            times.append(datetime.fromtimestamp(ctime).strftime("%Y-%m-%d %H:%M:%S"))
        except:
            continue
    return names, rolls, times

# ----------- Flask Routes -----------

@app.route('/')
def home():
    if 'username' not in session:
        return redirect(url_for('admin_login'))
    
    # Get today's attendance
    df = extract_attendance()
    tnames, trolls, ttimes = df['Name'].tolist(), df['Roll'].tolist(), df['Time'].tolist()
    
    # Get registered users
    rnames, rrolls, rtimes = get_registered_users()
    
    return render_template('home.html',
                           today_names=tnames,
                           today_rolls=trolls,
                           today_times=ttimes,
                           today_l=len(tnames),
                           reg_names=rnames,
                           reg_rolls=rrolls,
                           reg_times=rtimes,
                           reg_l=len(rnames),
                           totalreg=len(rnames),
                           datetoday2=date.today().strftime("%d-%B-%Y"),
                           mess=MESSAGE)

@app.route('/adminlogin', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        # Basic validation
        if not username or not password:
            flash('Please enter both username and password', 'danger')
            return render_template('adminlogin.html')
        
        # Authentication
        if username == 'admin123' and password == '1234':
            session['username'] = 'admin123'
            return redirect(url_for('home'))
        else:
            flash('Invalid credentials', 'danger')
    
    return render_template('adminlogin.html')

@app.route('/logout')
def logout():
    session.pop('username', None)
    flash('You have been logged out', 'info')
    return redirect(url_for('admin_login'))

@app.route('/admin')
def admin():
    if 'username' not in session:
        return redirect(url_for('admin_login'))
    
    # Get registered users data
    rnames, rrolls, rtimes = get_registered_users()
    
    return render_template('admin.html',
                         reg_names=rnames,
                         reg_rolls=rrolls,
                         reg_times=rtimes,
                         reg_l=len(rnames),
                         totalreg=len(rnames))

@app.route("/forgot-password")
def forgot_password():
    return render_template("sign.html")

@app.route('/start')
def start():
    global MESSAGE
    if not os.path.exists(MODEL_PATH):
        MESSAGE = "Model not trained. Register a face first!"
        flash(MESSAGE, 'warning')
        return redirect(url_for('home'))
    
    # Video capture setup with optimized settings
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        MESSAGE = "Error: Could not open camera"
        flash(MESSAGE, 'danger')
        return redirect(url_for('home'))
    
    # Set optimal camera parameters
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    cap.set(cv2.CAP_PROP_FPS, 30)
    cap.set(cv2.CAP_PROP_AUTOFOCUS, 0)  # Disable autofocus for better performance
    cap.set(cv2.CAP_PROP_FOCUS, 30)      # Set fixed focus
    
    marked_names = set()
    window_name = "Face Attendance System"
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
    
    # Performance optimization
    prev_time = time.time()
    frame_counter = 0
    skip_frames = 1  # Process every other frame
    last_valid_faces = []
    last_valid_boxes = []
    last_valid_names = []
    
    while True:
        ret, frame = cap.read()
        if not ret: 
            break
            
        # Check if window is still open
        if cv2.getWindowProperty(window_name, cv2.WND_PROP_VISIBLE) < 1:
            break
            
        # Performance optimization: Skip some frames
        frame_counter += 1
        if frame_counter % skip_frames != 0:
            # Display previous results if available
            if last_valid_faces:
                for (x, y, w, h), name in zip(last_valid_boxes, last_valid_names):
                    color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)
                    cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
                    cv2.putText(frame, name, (x, y - 10), 
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)
            
            # Display FPS
            curr_time = time.time()
            fps = 1 / (curr_time - prev_time)
            prev_time = curr_time
            cv2.putText(frame, f"FPS: {int(fps)}", (10, 30), 
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            cv2.imshow(window_name, frame)
            
            # Check for quit
            key = cv2.waitKey(1) & 0xFF
            if key == ord('q'):
                break
            continue
            
        # Process this frame
        faces, boxes = extract_faces_mtcnn(frame)
        current_names = []
        
        # Process faces in parallel
        def process_face(face):
            return recognize_face(face)
        
        # Submit face recognition tasks to thread pool
        future_to_face = {executor.submit(process_face, face): face for face in faces}
        for future in future_to_face:
            current_names.append(future.result())
        
        # Store results for next frame
        last_valid_faces = faces
        last_valid_boxes = boxes
        last_valid_names = current_names
        
        # Draw results on frame
        for face, (x, y, w, h), name in zip(faces, boxes, current_names):
            # Skip small or blurry faces
            if face.shape[0] < 40 or face.shape[1] < 40 or is_blurry(face):
                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 0, 255), 2)
                cv2.putText(frame, "Blurry/Small Face", (x, y-10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
                continue
                
            color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)
            cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
            cv2.putText(frame, name, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)
            
            # Add attendance when 'a' is pressed
            if name != "Unknown" and name not in marked_names:
                key = cv2.waitKey(1) & 0xFF
                if key == ord('a'):
                    add_attendance(name)
                    marked_names.add(name)
                    cv2.putText(frame, "ATTENDANCE MARKED", (x, y-40),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
        
        # Calculate and display FPS
        curr_time = time.time()
        fps = 1 / (curr_time - prev_time)
        prev_time = curr_time
        cv2.putText(frame, f"FPS: {int(fps)}", (10, 30), 
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
        
        cv2.imshow(window_name, frame)
        
        # Exit on 'q' press
        key = cv2.waitKey(1) & 0xFF
        if key == ord('q'):
            break
    
    # Cleanup
    cap.release()
    cv2.destroyAllWindows()
    return redirect(url_for('home'))

@app.route('/add', methods=['POST'])
def add():
    username = request.form.get('newusername', '').strip()
    userid = request.form.get('newuserid', '').strip()
    
    # Validate inputs
    if not username or not userid:
        flash('Please provide both name and ID', 'danger')
        return redirect(url_for('home'))
    
    folder = os.path.join(FACES_DIR, f"{username}_{userid}")
    os.makedirs(folder, exist_ok=True)
    
    # Video capture setup with optimized settings
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        flash('Error: Could not open camera', 'danger')
        return redirect(url_for('home'))
    
    # Set optimal camera parameters
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    cap.set(cv2.CAP_PROP_FPS, 30)
    cap.set(cv2.CAP_PROP_AUTOFOCUS, 0)  # Disable autofocus
    cap.set(cv2.CAP_PROP_FOCUS, 30)      # Set fixed focus
    
    count = 0
    window_name = "Register New User"
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
    
    # Performance optimization
    prev_time = time.time()
    frame_counter = 0
    skip_frames = 1  # Process every frame for registration
    
    while count < 20:  # Capture 20 good images
        ret, frame = cap.read()
        if not ret: 
            break
            
        if cv2.getWindowProperty(window_name, cv2.WND_PROP_VISIBLE) < 1:
            break
            
        # Performance optimization: Skip some frames
        frame_counter += 1
        if frame_counter % skip_frames != 0:
            cv2.imshow(window_name, frame)
            key = cv2.waitKey(1) & 0xFF
            if key == ord('q'):
                break
            continue
            
        # Face detection
        faces, boxes = extract_faces_mtcnn(frame)
        face_detected = False
        
        for face, (x, y, w, h) in zip(faces, boxes):
            # Skip small or blurry faces
            if face.shape[0] < 40 or face.shape[1] < 40 or is_blurry(face):
                cv2.putText(frame, "Blurry/Small", (x, y - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
                continue
                
            # Save good quality face
            count += 1
            img_path = os.path.join(folder, f"{username}_{count}.jpg")
            cv2.imwrite(img_path, face)
            face_detected = True
            
            # Visual feedback
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
            cv2.putText(frame, f"Saved: {count}/20", (x, y - 40),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            
            # Add slight delay between captures for better quality
            time.sleep(0.1)
            break  # Only capture one face per frame
        
        # Show progress and FPS
        curr_time = time.time()
        fps = 1 / (curr_time - prev_time)
        prev_time = curr_time
        cv2.putText(frame, f"Captured: {count}/20", (10, 30), 
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
        cv2.putText(frame, f"FPS: {int(fps)}", (10, 60), 
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
        
        # Provide guidance if no face detected
        if not face_detected and count < 20:
            cv2.putText(frame, "Position face clearly in frame", (10, 90),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
        
        cv2.imshow(window_name, frame)
        
        # Exit on 'q' press
        key = cv2.waitKey(1) & 0xFF
        if key == ord('q'):
            break
        
        # Break early if we have enough images
        if count >= 20:
            break
    
    # Cleanup
    cap.release()
    cv2.destroyAllWindows()
     
    # Train model with new data
    train_model()
    
    flash(f'User {username} added successfully! Model is training in background.', 'success')
    return redirect(url_for('home'))

# ----------- Run Flask App -----------
if __name__ == '__main__':
    app.run(debug=True, port=1000)
